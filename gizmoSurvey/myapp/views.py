from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.db import models
from django.utils import timezone
from .models import UserProfile, Survey, Question, SurveyResponse, Answer, Section
from .forms import UserRegistrationForm, SurveyForm, QuestionForm, SurveyResponseForm, SectionForm, QuestionBulkForm, SurveySettingsForm, AssignmentForm, SectionBulkForm
import json
import csv
import io
from datetime import datetime


def home(request):
    """Home page with login/signup options"""
    if request.user.is_authenticated:
        try:
            profile = UserProfile.objects.get(user=request.user)
            # Check if user profile is active
            if not profile.is_active:
                logout(request)
                messages.error(request, 'Your account has been deactivated. Please contact your teacher.')
                return render(request, 'myapp/home.html')
            
            if profile.role == 'student':
                return redirect('student_dashboard')
            else:
                return redirect('teacher_dashboard')
        except UserProfile.DoesNotExist:
            logout(request)
            messages.error(request, 'User profile not found. Please contact your teacher.')
            return render(request, 'myapp/home.html')
    return render(request, 'myapp/home.html')


def custom_login(request):
    """Custom login view with show password functionality"""
    if request.user.is_authenticated:
        try:
            profile = UserProfile.objects.get(user=request.user)
            # Check if user profile is active
            if not profile.is_active:
                logout(request)
                messages.error(request, 'Your account has been deactivated. Please contact your teacher.')
                return redirect('home')
            
            if profile.role == 'student':
                return redirect('student_dashboard')
            else:
                return redirect('teacher_dashboard')
        except UserProfile.DoesNotExist:
            logout(request)
            messages.error(request, 'User profile not found. Please contact your teacher.')
            return redirect('home')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        if username and password:
            user = authenticate(request, username=username, password=password)
            if user is not None:
                # Check if user profile exists and is active
                try:
                    profile = UserProfile.objects.get(user=user)
                    if not profile.is_active:
                        messages.error(request, 'Your account has been deactivated. Please contact your teacher.')
                        return render(request, 'myapp/login.html')
                except UserProfile.DoesNotExist:
                    messages.error(request, 'User profile not found. Please contact your teacher.')
                    return render(request, 'myapp/login.html')
                
                login(request, user)
                messages.success(request, f'Welcome back, {user.get_full_name() or user.username}!')
                if profile.role == 'student':
                    return redirect('student_dashboard')
                else:
                    return redirect('teacher_dashboard')
            else:
                messages.error(request, 'Invalid username or password.')
        else:
            messages.error(request, 'Please fill in both username and password.')
    
    return render(request, 'myapp/login.html')


def custom_logout(request):
    """Custom logout view"""
    logout(request)
    messages.success(request, 'You have been successfully logged out.')
    return redirect('home')


def register(request):
    """User registration with role selection"""
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password1')
            user = authenticate(username=username, password=password)
            login(request, user)
            messages.success(request, 'Registration successful!')
            return redirect('home')
    else:
        form = UserRegistrationForm()
    return render(request, 'myapp/register.html', {'form': form})


@login_required
def student_dashboard(request):
    """Student dashboard showing assigned surveys"""
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'student':
        messages.error(request, 'Access denied. Student access required.')
        return redirect('home')
    
    # Check if student is active
    if not profile.is_active:
        messages.error(request, 'Your account has been deactivated. Please contact your teacher.')
        return redirect('home')
    
    # Get surveys assigned to student's section
    assigned_surveys = Survey.objects.filter(
        sections=profile.section,
        is_active=True
    ).order_by('-created_at')
    
    # Check which surveys student has already completed (current version)
    completed_surveys = []
    for survey in assigned_surveys:
        latest_response = SurveyResponse.objects.filter(
            survey=survey, 
            student=request.user
        ).order_by('-submitted_at').first()
        
        if latest_response and latest_response.survey_version >= survey.version:
            completed_surveys.append(survey.id)
    
    # Check which surveys need to be retaken (outdated responses)
    surveys_to_retake = []
    for survey in assigned_surveys:
        latest_response = SurveyResponse.objects.filter(
            survey=survey, 
            student=request.user
        ).order_by('-submitted_at').first()
        
        if latest_response and latest_response.survey_version < survey.version:
            surveys_to_retake.append(survey.id)
    
    # Calculate pending surveys (assigned - completed - retake)
    pending_count = assigned_surveys.count() - len(completed_surveys) - len(surveys_to_retake)
    
    context = {
        'assigned_surveys': assigned_surveys,
        'completed_surveys': completed_surveys,
        'surveys_to_retake': surveys_to_retake,
        'pending_count': pending_count,
        'profile': profile,
    }
    return render(request, 'myapp/student_dashboard.html', context)


@login_required
def take_survey(request, survey_id):
    """Student takes a survey"""
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'student':
        messages.error(request, 'Access denied. Student access required.')
        return redirect('home')
    
    # Check if student is active
    if not profile.is_active:
        messages.error(request, 'Your account has been deactivated. Please contact your teacher.')
        return redirect('home')
    
    survey = get_object_or_404(Survey, id=survey_id)
    
    # Check if survey is assigned to student's section
    if profile.section not in survey.sections.all():
        messages.error(request, 'This survey is not assigned to your section.')
        return redirect('student_dashboard')
    
    # Check if survey is still open
    if not survey.is_open:
        messages.error(request, 'This survey is no longer accepting responses.')
        return redirect('student_dashboard')
    
    # Check if student has already completed the current version of this survey
    latest_response = SurveyResponse.objects.filter(survey=survey, student=request.user).order_by('-submitted_at').first()
    if latest_response and latest_response.survey_version >= survey.version:
        messages.info(request, 'You have already completed this survey.')
        return redirect('student_dashboard')
    
    if request.method == 'POST':
        form = SurveyResponseForm(survey, request.POST)
        if form.is_valid():
            # Create survey response
            response = SurveyResponse.objects.create(
                survey=survey,
                student=request.user,
                survey_version=survey.version
            )
            
            # Save answers
            for question in survey.questions.filter(is_active=True):
                field_name = f'question_{question.id}'
                if field_name in form.cleaned_data:
                    answer_value = form.cleaned_data[field_name]
                    
                    answer = Answer.objects.create(
                        response=response,
                        question=question
                    )
                    
                    if question.question_type in ['multiple_choice', 'likert_scale']:
                        answer.answer_choice = answer_value
                    elif question.question_type in ['short_answer', 'long_answer']:
                        answer.answer_text = answer_value
                    
                    answer.save()
            
            messages.success(request, 'Survey submitted successfully!')
            return redirect('student_dashboard')
    else:
        form = SurveyResponseForm(survey)
    
    context = {
        'survey': survey,
        'form': form,
        'profile': profile,
    }
    return render(request, 'myapp/take_survey.html', context)


@login_required
def student_history(request):
    """Student's survey response history"""
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'student':
        messages.error(request, 'Access denied. Student access required.')
        return redirect('home')
    
    # Check if student is active
    if not profile.is_active:
        messages.error(request, 'Your account has been deactivated. Please contact your teacher.')
        return redirect('home')
    
    responses = SurveyResponse.objects.filter(student=request.user).order_by('-submitted_at')
    
    # Add original question count for each response
    for response in responses:
        # Count questions that existed when this response was submitted
        # We'll use the number of answers as the original question count
        response.original_question_count = response.answers.count()
    
    # Pagination
    paginator = Paginator(responses, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'profile': profile,
    }
    return render(request, 'myapp/student_history.html', context)


@login_required
def survey_list(request):
    """Teacher survey list with management options"""
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'teacher':
        messages.error(request, 'Access denied. Teacher access required.')
        return redirect('home')
    
    # Get teacher's surveys with search and filtering
    surveys = Survey.objects.filter(created_by=request.user).order_by('-created_at')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        surveys = surveys.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter == 'active':
        surveys = surveys.filter(is_active=True)
    elif status_filter == 'inactive':
        surveys = surveys.filter(is_active=False)
    
    # Add statistics for each survey
    survey_data = []
    for survey in surveys:
        total_responses = survey.responses.count()
        current_version_responses = survey.responses.filter(survey_version=survey.version).count()
        outdated_version_responses = survey.responses.exclude(survey_version=survey.version).count()
        
        # Get assigned sections
        assigned_sections = survey.sections.filter(is_active=True)
        
        survey_data.append({
            'survey': survey,
            'total_responses': total_responses,
            'current_version_responses': current_version_responses,
            'outdated_version_responses': outdated_version_responses,
            'assigned_sections': assigned_sections,
            'question_count': survey.questions.filter(is_active=True).count(),
        })
    
    # Pagination
    paginator = Paginator(survey_data, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'profile': profile,
    }
    return render(request, 'myapp/survey_list.html', context)


@login_required
def teacher_dashboard(request):
    """Teacher dashboard with survey management and analytics"""
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'teacher':
        messages.error(request, 'Access denied. Teacher access required.')
        return redirect('home')
    
    # Get teacher's surveys
    surveys = Survey.objects.filter(created_by=request.user).order_by('-created_at')
    
    # Get response statistics with analytics data
    survey_stats = []
    active_surveys_count = 0
    total_responses_count = 0
    
    for survey in surveys:
        total_responses = survey.responses.count()
        total_responses_count += total_responses
        
        # Calculate version-specific counts
        current_version_responses = survey.responses.filter(survey_version=survey.version).count()
        outdated_version_responses = survey.responses.exclude(survey_version=survey.version).count()
        
        # Get analytics data for this survey
        analytics_data = get_survey_analytics_data(survey)
        
        # Count active surveys
        if survey.is_open:
            active_surveys_count += 1
        
        survey_stats.append({
            'survey': survey,
            'total_responses': total_responses,
            'current_version_responses': current_version_responses,
            'outdated_version_responses': outdated_version_responses,
            'is_open': survey.is_open,
            'analytics_data': analytics_data,
            'active_questions_count': survey.questions.filter(is_active=True).count(),
        })
    
    # Get dashboard analytics data
    dashboard_analytics = get_dashboard_analytics_data(request.user)
    
    context = {
        'survey_stats': survey_stats,
        'active_surveys_count': active_surveys_count,
        'total_responses_count': total_responses_count,
        'profile': profile,
        'dashboard_analytics': dashboard_analytics,
        'surveys_for_filter': surveys,
    }
    return render(request, 'myapp/teacher_dashboard.html', context)


@login_required
def create_survey(request):
    """Create a new survey"""
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'teacher':
        messages.error(request, 'Access denied. Teacher access required.')
        return redirect('home')
    
    if request.method == 'POST':
        form = SurveyForm(request.POST)
        if form.is_valid():
            survey = form.save(commit=False)
            survey.created_by = request.user
            survey.save()
            form.save_m2m()  # Save many-to-many relationships
            messages.success(request, 'Survey created successfully!')
            return redirect('edit_survey', survey_id=survey.id)
    else:
        form = SurveyForm()
    
    context = {
        'form': form,
        'profile': profile,
    }
    return render(request, 'myapp/create_survey.html', context)


@login_required
def edit_survey(request, survey_id):
    """Edit survey and manage questions"""
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'teacher':
        messages.error(request, 'Access denied. Teacher access required.')
        return redirect('home')
    
    survey = get_object_or_404(Survey, id=survey_id, created_by=request.user)
    
    if request.method == 'POST':
        form = SurveyForm(request.POST, instance=survey)
        if form.is_valid():
            # Check if content fields changed (not administrative fields)
            content_fields = ['title', 'description']
            content_changed = any(
                form.cleaned_data.get(field) != getattr(survey, field)
                for field in content_fields
            )
            
            # Only increment version if content changed and survey has responses
            if content_changed and survey.responses.exists():
                survey.version += 1
                messages.success(request, f'Survey updated successfully! Version incremented to {survey.version}. Students will need to retake the survey.')
            else:
                messages.success(request, 'Survey updated successfully!')
            
            form.save()
            return redirect('edit_survey', survey_id=survey.id)
    else:
        form = SurveyForm(instance=survey)
    
    # Get questions for this survey
    questions = survey.questions.filter(is_active=True).order_by('order')
    inactive_questions = survey.questions.filter(is_active=False).order_by('order')
    
    context = {
        'survey': survey,
        'form': form,
        'questions': questions,
        'inactive_questions': inactive_questions,
        'profile': profile,
    }
    return render(request, 'myapp/edit_survey.html', context)


@login_required
def add_question(request, survey_id):
    """Add a question to a survey or batch save multiple questions"""
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'teacher':
        messages.error(request, 'Access denied. Teacher access required.')
        return redirect('home')
    
    survey = get_object_or_404(Survey, id=survey_id, created_by=request.user)
    
    if request.method == 'POST':
        # Check if this is a batch save request
        if request.content_type == 'application/json':
            try:
                data = json.loads(request.body)
                if data.get('batch_save') and 'questions' in data:
                    return handle_batch_save(request, survey, data['questions'])
            except json.JSONDecodeError:
                return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
        
        # Regular single question save
        form = QuestionForm(request.POST)
        if form.is_valid():
            question = form.save(commit=False)
            question.survey = survey
            question.save()
            
            # Increment survey version if it has responses
            if survey.responses.exists():
                survey.version += 1
                survey.save()
                messages.success(request, f'Question added successfully! Survey version incremented to {survey.version}. Students will need to retake the survey.')
            else:
                messages.success(request, 'Question added successfully!')
            
            return redirect('edit_survey', survey_id=survey.id)
    else:
        form = QuestionForm()
    
    context = {
        'survey': survey,
        'form': form,
        'profile': profile,
    }
    return render(request, 'myapp/add_question.html', context)


def handle_batch_save(request, survey, questions_data):
    """Handle batch saving of multiple questions"""
    try:
        created_questions = []
        version_incremented = False
        
        for question_data in questions_data:
            # Create question object
            question = Question(
                survey=survey,
                question_text=question_data.get('question_text', ''),
                question_type=question_data.get('question_type', 'short_answer'),
                is_required=question_data.get('is_required', True),
                order=question_data.get('order', 1),
                options=question_data.get('options', []),
                likert_min=question_data.get('likert_min', 1),
                likert_max=question_data.get('likert_max', 5),
                likert_labels=question_data.get('likert_labels', [])
            )
            
            # Validate the question
            if not question.question_text.strip():
                return JsonResponse({'success': False, 'error': 'Question text cannot be empty'})
            
            question.save()
            created_questions.append(question)
        
        # Increment survey version if it has responses
        if survey.responses.exists():
            survey.version += 1
            survey.save()
            version_incremented = True
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully saved {len(created_questions)} questions!',
            'version_incremented': version_incremented,
            'new_version': survey.version if version_incremented else None
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def edit_question(request, question_id):
    """Edit a question"""
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'teacher':
        messages.error(request, 'Access denied. Teacher access required.')
        return redirect('home')
    
    question = get_object_or_404(Question, id=question_id, survey__created_by=request.user)
    
    if request.method == 'POST':
        form = QuestionForm(request.POST, instance=question)
        if form.is_valid():
            form.save()
            
            # Increment survey version if it has responses
            if question.survey.responses.exists():
                question.survey.version += 1
                question.survey.save()
                messages.success(request, f'Question updated successfully! Survey version incremented to {question.survey.version}. Students will need to retake the survey.')
            else:
                messages.success(request, 'Question updated successfully!')
            
            return redirect('edit_survey', survey_id=question.survey.id)
    else:
        form = QuestionForm(instance=question)
    
    context = {
        'question': question,
        'form': form,
        'profile': profile,
    }
    return render(request, 'myapp/edit_question.html', context)


@login_required
def delete_question(request, question_id):
    """Soft delete a question (mark as inactive)"""
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'teacher':
        messages.error(request, 'Access denied. Teacher access required.')
        return redirect('home')
    
    question = get_object_or_404(Question, id=question_id, survey__created_by=request.user)
    survey = question.survey
    survey_id = survey.id
    
    # Increment survey version if it has responses
    if survey.responses.exists():
        survey.version += 1
        survey.save()
        messages.success(request, f'Question deactivated successfully! Survey version incremented to {survey.version}. Students will need to retake the survey.')
    else:
        messages.success(request, 'Question deactivated successfully!')
    
    # Soft delete: mark as inactive instead of deleting
    question.is_active = False
    question.save()
    return redirect('edit_survey', survey_id=survey_id)


@login_required
def restore_question(request, question_id):
    """Restore a soft-deleted question (mark as active)"""
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'teacher':
        messages.error(request, 'Access denied. Teacher access required.')
        return redirect('home')
    
    question = get_object_or_404(Question, id=question_id, survey__created_by=request.user)
    survey = question.survey
    survey_id = survey.id
    
    # Increment survey version if it has responses
    if survey.responses.exists():
        survey.version += 1
        survey.save()
        messages.success(request, f'Question restored successfully! Survey version incremented to {survey.version}. Students will need to retake the survey.')
    else:
        messages.success(request, 'Question restored successfully!')
    
    # Restore: mark as active
    question.is_active = True
    question.save()
    return redirect('edit_survey', survey_id=survey_id)


@login_required
def survey_responses(request, survey_id):
    """View survey responses grouped by version"""
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'teacher':
        messages.error(request, 'Access denied. Teacher access required.')
        return redirect('home')
    
    survey = get_object_or_404(Survey, id=survey_id, created_by=request.user)
    responses = SurveyResponse.objects.filter(survey=survey).order_by('-submitted_at')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        responses = responses.filter(
            Q(student__first_name__icontains=search_query) |
            Q(student__last_name__icontains=search_query) |
            Q(student__username__icontains=search_query)
        )
    
    # Group responses by version
    responses_by_version = {}
    for response in responses:
        version = response.survey_version
        if version not in responses_by_version:
            responses_by_version[version] = []
        responses_by_version[version].append(response)
    
    # Sort versions in descending order (newest first)
    sorted_versions = sorted(responses_by_version.keys(), reverse=True)
    
    # Add original question count for each response
    for version_responses in responses_by_version.values():
        for response in version_responses:
            response.original_question_count = response.answers.count()
    
    # Calculate version counts
    current_version_count = 0
    outdated_version_counts = {}
    
    for version, version_responses in responses_by_version.items():
        if version == survey.version:
            current_version_count = len(version_responses)
        else:
            outdated_version_counts[version] = len(version_responses)
    
    context = {
        'survey': survey,
        'responses_by_version': responses_by_version,
        'sorted_versions': sorted_versions,
        'current_version_count': current_version_count,
        'outdated_version_counts': outdated_version_counts,
        'search_query': search_query,
        'profile': profile,
    }
    return render(request, 'myapp/survey_responses.html', context)


@login_required
def view_response(request, response_id):
    """View individual response details"""
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'teacher':
        messages.error(request, 'Access denied. Teacher access required.')
        return redirect('home')
    
    response = get_object_or_404(SurveyResponse, id=response_id, survey__created_by=request.user)
    answers = response.answers.all().order_by('question__order')
    
    context = {
        'response': response,
        'answers': answers,
        'profile': profile,
    }
    return render(request, 'myapp/view_response.html', context)


@login_required
def survey_analytics(request, survey_id):
    """Enhanced survey analytics and visualizations with real-time data"""
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'teacher':
        messages.error(request, 'Access denied. Teacher access required.')
        return redirect('home')
    
    survey = get_object_or_404(Survey, id=survey_id, created_by=request.user)
    
    # Get comprehensive analytics data
    analytics_data = get_survey_analytics_data(survey)
    
    # Calculate survey statistics
    total_responses = survey.responses.count()
    total_questions = survey.questions.count()
    
    # Calculate completion rate by section
    section_stats = []
    for section in survey.sections.all():
        students_in_section = UserProfile.objects.filter(section=section, role='student', is_active=True).count()
        responses_from_section = SurveyResponse.objects.filter(
            survey=survey,
            student__userprofile__section=section
        ).count()
        
        completion_rate = (responses_from_section / students_in_section * 100) if students_in_section > 0 else 0
        
        section_stats.append({
            'section': section,
            'total_students': students_in_section,
            'responses_received': responses_from_section,
            'completion_rate': round(completion_rate, 1),
        })
    
    context = {
        'survey': survey,
        'analytics_data': analytics_data,
        'total_responses': total_responses,
        'total_questions': total_questions,
        'section_stats': section_stats,
        'profile': profile,
    }
    return render(request, 'myapp/survey_analytics.html', context)


def get_survey_analytics_data(survey):
    """Helper function to get comprehensive analytics data for a survey"""
    analytics_data = []
    
    for question in survey.questions.all():
        question_data = {
            'question': question,
            'type': question.question_type,
            'responses': [],
            'stats': {},
            'chart_data': {},
            'word_cloud_data': []
        }
        
        answers = Answer.objects.filter(response__survey=survey, question=question)
        
        if question.question_type == 'multiple_choice':
            # Count choices with percentages
            choice_counts = {}
            total_answers = answers.count()
            
            for answer in answers:
                choice = answer.answer_choice
                choice_counts[choice] = choice_counts.get(choice, 0) + 1
            
            # Calculate percentages
            choice_stats = {}
            for choice, count in choice_counts.items():
                percentage = (count / total_answers * 100) if total_answers > 0 else 0
                choice_stats[choice] = {
                    'count': count,
                    'percentage': round(percentage, 1)
                }
            
            question_data['stats'] = choice_stats
            question_data['chart_data'] = {
                'labels': list(choice_counts.keys()),
                'data': list(choice_counts.values()),
                'type': 'pie'
            }
        
        elif question.question_type == 'likert_scale':
            # Count scale values with percentages
            scale_counts = {}
            total_answers = answers.count()
            
            for answer in answers:
                value = answer.answer_choice
                scale_counts[value] = scale_counts.get(value, 0) + 1
            
            # Sort by scale value
            sorted_scales = sorted(scale_counts.items(), key=lambda x: int(x[0]) if x[0].isdigit() else x[0])
            
            scale_stats = {}
            for value, count in sorted_scales:
                percentage = (count / total_answers * 100) if total_answers > 0 else 0
                scale_stats[value] = {
                    'count': count,
                    'percentage': round(percentage, 1)
                }
            
            question_data['stats'] = scale_stats
            question_data['chart_data'] = {
                'labels': [item[0] for item in sorted_scales],
                'data': [item[1] for item in sorted_scales],
                'type': 'bar'
            }
        
        elif question.question_type in ['short_answer', 'long_answer']:
            # Collect text responses for word cloud
            text_responses = [answer.answer_text.strip() for answer in answers if answer.answer_text and answer.answer_text.strip()]
            question_data['responses'] = text_responses
            
            # Process text for word cloud
            word_frequency = process_text_for_wordcloud(text_responses)
            question_data['word_cloud_data'] = word_frequency
        
        analytics_data.append(question_data)
    
    return analytics_data


def process_text_for_wordcloud(text_responses):
    """Process text responses to create word frequency data for word clouds"""
    import re
    from collections import Counter
    
    # Combine all text responses
    all_text = ' '.join(text_responses).lower()
    
    # Remove common stop words and clean text
    stop_words = {
        'the', 'a', 'an', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for', 'of', 'with', 'by',
        'is', 'are', 'was', 'were', 'be', 'been', 'have', 'has', 'had', 'do', 'does', 'did',
        'will', 'would', 'could', 'should', 'may', 'might', 'must', 'can', 'this', 'that',
        'these', 'those', 'i', 'you', 'he', 'she', 'it', 'we', 'they', 'me', 'him', 'her',
        'us', 'them', 'my', 'your', 'his', 'her', 'its', 'our', 'their', 'very', 'really',
        'quite', 'just', 'only', 'also', 'too', 'so', 'as', 'if', 'when', 'where', 'why',
        'how', 'what', 'who', 'which', 'there', 'here', 'now', 'then', 'than', 'more',
        'most', 'some', 'any', 'all', 'both', 'each', 'every', 'no', 'not', 'yes'
    }
    
    # Extract words (alphanumeric characters only)
    words = re.findall(r'\b[a-zA-Z]{3,}\b', all_text)
    
    # Filter out stop words and count frequency
    filtered_words = [word for word in words if word not in stop_words]
    word_counts = Counter(filtered_words)
    
    # Return top 50 words with their frequencies
    return [{'text': word, 'weight': count} for word, count in word_counts.most_common(50)]


def get_filtered_survey_analytics_data(survey, responses_query, question_type_filter=None):
    """Get filtered analytics data for a survey based on responses query and question type filter"""
    analytics_data = []
    
    questions_query = survey.questions.all()
    if question_type_filter and question_type_filter != 'all':
        questions_query = questions_query.filter(question_type=question_type_filter)
    
    for question in questions_query:
        question_data = {
            'question': {
                'id': question.id,
                'text': question.question_text,
                'type': question.question_type,
                'version': survey.version,  # Use survey version instead of question version
                'is_required': question.is_required
            },
            'type': question.question_type,
            'responses': [],
            'stats': {},
            'chart_data': {},
            'word_cloud_data': []
        }
        
        # Get answers filtered by responses query
        answers = Answer.objects.filter(
            response__in=responses_query,
            question=question
        )
        
        # Get response version information
        response_versions = []
        for answer in answers:
            response_versions.append({
                'survey_version': answer.response.survey_version,
                'answer_text': answer.answer_text,
                'submitted_at': answer.response.submitted_at
            })
        
        question_data['responses'] = response_versions
        
        if question.question_type == 'multiple_choice':
            # Count choices with percentages
            choice_counts = {}
            total_answers = answers.count()
            
            for answer in answers:
                choice = answer.answer_choice
                choice_counts[choice] = choice_counts.get(choice, 0) + 1
            
            # Calculate percentages
            choice_stats = {}
            for choice, count in choice_counts.items():
                percentage = (count / total_answers * 100) if total_answers > 0 else 0
                choice_stats[choice] = {
                    'count': count,
                    'percentage': round(percentage, 1)
                }
            
            question_data['stats'] = choice_stats
            question_data['chart_data'] = {
                'labels': list(choice_counts.keys()),
                'data': list(choice_counts.values()),
                'type': 'pie'
            }
        
        elif question.question_type == 'likert_scale':
            # Count scale values with percentages
            scale_counts = {}
            total_answers = answers.count()
            
            for answer in answers:
                value = answer.answer_choice
                scale_counts[value] = scale_counts.get(value, 0) + 1
            
            # Sort by scale value
            sorted_scales = sorted(scale_counts.items(), key=lambda x: int(x[0]) if x[0].isdigit() else x[0])
            
            scale_stats = {}
            for value, count in sorted_scales:
                percentage = (count / total_answers * 100) if total_answers > 0 else 0
                scale_stats[value] = {
                    'count': count,
                    'percentage': round(percentage, 1)
                }
            
            question_data['stats'] = scale_stats
            question_data['chart_data'] = {
                'labels': [item[0] for item in sorted_scales],
                'data': [item[1] for item in sorted_scales],
                'type': 'bar'
            }
        
        elif question.question_type in ['short_answer', 'long_answer']:
            # Collect text responses for word cloud
            text_responses = [answer.answer_text.strip() for answer in answers if answer.answer_text and answer.answer_text.strip()]
            question_data['responses'] = text_responses
            
            # Process text for word cloud
            word_frequency = process_text_for_wordcloud(text_responses)
            question_data['word_cloud_data'] = word_frequency
        
        analytics_data.append(question_data)
    
    return analytics_data


def generate_response_timeline_data(responses_query):
    """Generate timeline data for response trends over the last 30 days"""
    from django.utils import timezone
    from datetime import timedelta
    from django.db.models import Count
    
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)
    
    # Get daily response counts
    daily_responses = responses_query.filter(
        submitted_at__date__gte=start_date,
        submitted_at__date__lte=end_date
    ).extra(
        select={'day': 'date(submitted_at)'}
    ).values('day').annotate(count=Count('id')).order_by('day')
    
    # Create a complete timeline with all days
    timeline_data = []
    current_date = start_date
    
    # Create a dictionary for quick lookup
    response_counts = {item['day']: item['count'] for item in daily_responses}
    
    while current_date <= end_date:
        count = response_counts.get(current_date, 0)
        timeline_data.append({
            'date': current_date.strftime('%Y-%m-%d'),
            'date_formatted': current_date.strftime('%m/%d'),
            'response_count': count
        })
        current_date += timedelta(days=1)
    
    return timeline_data


def generate_analytics_insights(survey, total_responses, section_stats, analytics_data):
    """Generate AI-powered insights based on survey data"""
    insights = []
    
    # Response rate analysis
    if total_responses > 0:
        avg_completion = sum(stat['completion_rate'] for stat in section_stats) / len(section_stats) if section_stats else 0
        
        if avg_completion >= 80:
            insights.append({
                'type': 'success',
                'title': 'Excellent Response Rate',
                'message': f'Your survey has achieved an {avg_completion:.1f}% average completion rate across all sections. This indicates strong student engagement.',
                'icon': 'check-circle'
            })
        elif avg_completion >= 50:
            insights.append({
                'type': 'warning',
                'title': 'Moderate Response Rate',
                'message': f'Your survey has a {avg_completion:.1f}% average completion rate. Consider sending reminders to improve participation.',
                'icon': 'exclamation-triangle'
            })
        else:
            insights.append({
                'type': 'error',
                'title': 'Low Response Rate',
                'message': f'Your survey has a {avg_completion:.1f}% average completion rate. Consider extending the deadline or sending multiple reminders.',
                'icon': 'exclamation-circle'
            })
    else:
        insights.append({
            'type': 'info',
            'title': 'No Responses Yet',
            'message': 'No responses have been received yet. Consider sending reminders to students or extending the deadline.',
            'icon': 'info-circle'
        })
    
    # Question analysis
    if analytics_data:
        # Analyze question types
        question_types = {}
        for data in analytics_data:
            q_type = data['question']['type']  # Access the 'type' key from the dictionary
            question_types[q_type] = question_types.get(q_type, 0) + 1
        
        if len(question_types) > 1:
            insights.append({
                'type': 'info',
                'title': 'Question Diversity',
                'message': f'Your survey includes {len(question_types)} different question types, providing a good mix of quantitative and qualitative data.',
                'icon': 'lightbulb'
            })
        
        # Analyze text responses
        text_questions = [data for data in analytics_data if data['question']['type'] in ['short_answer', 'long_answer']]
        if text_questions:
            total_text_responses = sum(len(data['responses']) for data in text_questions)
            if total_text_responses > 0:
                insights.append({
                    'type': 'success',
                    'title': 'Rich Text Data',
                    'message': f'You have collected {total_text_responses} text responses across {len(text_questions)} open-ended questions, providing valuable qualitative insights.',
                    'icon': 'chat-text'
                })
    
    # Survey status insights
    if survey.is_open:
        if survey.due_date:
            from django.utils import timezone
            days_remaining = (survey.due_date.date() - timezone.now().date()).days
            if days_remaining <= 3:
                insights.append({
                    'type': 'warning',
                    'title': 'Deadline Approaching',
                    'message': f'Your survey deadline is in {days_remaining} days. Consider sending final reminders to students.',
                    'icon': 'clock'
                })
    else:
        insights.append({
            'type': 'info',
            'title': 'Survey Closed',
            'message': 'Your survey is currently closed. Consider creating a new version for future data collection.',
            'icon': 'pause-circle'
        })
    
    return insights


def generate_version_timeline_data(survey, responses_query=None):
    """Generate timeline data comparing current vs outdated versions"""
    from django.utils import timezone
    from datetime import timedelta
    from django.db.models import Count
    
    # Use provided responses_query or get all responses
    if responses_query is None:
        responses_query = SurveyResponse.objects.filter(survey=survey)
    
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)
    
    # Get daily response counts for current version
    current_daily = responses_query.filter(
        survey_version=survey.version,
        submitted_at__date__gte=start_date,
        submitted_at__date__lte=end_date
    ).extra(
        select={'day': 'date(submitted_at)'}
    ).values('day').annotate(count=Count('id')).order_by('day')
    
    # Get daily response counts for outdated versions
    outdated_daily = responses_query.filter(
        submitted_at__date__gte=start_date,
        submitted_at__date__lte=end_date
    ).exclude(survey_version=survey.version).extra(
        select={'day': 'date(submitted_at)'}
    ).values('day').annotate(count=Count('id')).order_by('day')
    
    # Create complete timeline
    timeline_data = []
    current_date = start_date
    
    # Create dictionaries for quick lookup
    current_counts = {item['day']: item['count'] for item in current_daily}
    outdated_counts = {item['day']: item['count'] for item in outdated_daily}
    
    while current_date <= end_date:
        current_count = current_counts.get(current_date, 0)
        outdated_count = outdated_counts.get(current_date, 0)
        
        timeline_data.append({
            'date': current_date.strftime('%Y-%m-%d'),
            'date_formatted': current_date.strftime('%m/%d'),
            'current': current_count,
            'outdated': outdated_count
        })
        current_date += timedelta(days=1)
    
    return {
        'timeline': timeline_data,
        'comparison': {
            'current': responses_query.filter(survey_version=survey.version).count(),
            'outdated': responses_query.exclude(survey_version=survey.version).count()
        }
    }


@login_required
def student_response_details(request, response_id):
    """Get detailed response information for AJAX requests"""
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'student':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    try:
        response = SurveyResponse.objects.get(id=response_id, student=request.user)
        answers = response.answers.all().order_by('question__order')
        
        response_data = {
            'survey_title': response.survey.title,
            'submitted_at': response.submitted_at.strftime('%B %d, %Y at %I:%M %p'),
            'is_complete': response.is_complete,
            'answers': []
        }
        
        for answer in answers:
            answer_data = {
                'question_text': answer.question.question_text,
                'question_type': answer.question.question_type,
                'answer_value': '',
                'question_options': []
            }
            
            if answer.question.question_type == 'multiple_choice':
                answer_data['answer_value'] = answer.answer_choice
                answer_data['question_options'] = answer.question.options
            elif answer.question.question_type == 'likert_scale':
                answer_data['answer_value'] = answer.answer_choice
                answer_data['question_options'] = answer.question.likert_labels
                answer_data['likert_min'] = answer.question.likert_min
                answer_data['likert_max'] = answer.question.likert_max
            elif answer.question.question_type in ['short_answer', 'long_answer']:
                answer_data['answer_value'] = answer.answer_text
            
            response_data['answers'].append(answer_data)
        
        return JsonResponse(response_data)
    
    except SurveyResponse.DoesNotExist:
        return JsonResponse({'error': 'Response not found'}, status=404)


@login_required
def manage_sections(request):
    """Manage sections"""
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'teacher':
        messages.error(request, 'Access denied. Teacher access required.')
        return redirect('home')
    
    sections = Section.objects.all().order_by('-is_active', 'name')  # Active sections first, then inactive
    inactive_sections = Section.objects.filter(is_active=False).order_by('name')
    
    if request.method == 'POST':
        form = SectionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Section created successfully!')
            return redirect('manage_sections')
    else:
        form = SectionForm()
    
    context = {
        'sections': sections,
        'inactive_sections': inactive_sections,
        'form': form,
        'profile': profile,
    }
    return render(request, 'myapp/manage_sections.html', context)


# Enhanced CRUD Views for Survey Builder

@login_required
def question_bulk_operations(request, survey_id):
    """Handle bulk operations on questions"""
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'teacher':
        messages.error(request, 'Access denied. Teacher access required.')
        return redirect('home')
    
    survey = get_object_or_404(Survey, id=survey_id, created_by=request.user)
    
    if request.method == 'POST':
        form = QuestionBulkForm(request.POST)
        if form.is_valid():
            action = form.cleaned_data['action']
            question_ids = form.cleaned_data['question_ids'].split(',')
            questions = Question.objects.filter(id__in=question_ids, survey=survey)
            
            # Check if survey has responses to determine if version should be incremented
            version_incremented = False
            if survey.responses.exists():
                survey.version += 1
                survey.save()
                version_incremented = True
            
            if action == 'delete':
                count = questions.count()
                # Soft delete: mark as inactive instead of deleting
                questions.update(is_active=False)
                if version_incremented:
                    messages.success(request, f'{count} questions deactivated successfully! Survey version incremented to {survey.version}. Students will need to retake the survey.')
                else:
                    messages.success(request, f'{count} questions deactivated successfully!')
            
            elif action == 'reorder':
                new_order = form.cleaned_data['new_order']
                if new_order:
                    order_numbers = [int(x.strip()) for x in new_order.split(',') if x.strip()]
                    for i, question in enumerate(questions):
                        if i < len(order_numbers):
                            question.order = order_numbers[i]
                            question.save()
                    if version_incremented:
                        messages.success(request, f'Questions reordered successfully! Survey version incremented to {survey.version}. Students will need to retake the survey.')
                    else:
                        messages.success(request, 'Questions reordered successfully!')
            
            elif action == 'toggle_required':
                for question in questions:
                    question.is_required = not question.is_required
                    question.save()
                if version_incremented:
                    messages.success(request, f'Required status toggled successfully! Survey version incremented to {survey.version}. Students will need to retake the survey.')
                else:
                    messages.success(request, 'Required status toggled successfully!')
            
            elif action == 'change_type':
                new_type = form.cleaned_data['new_type']
                if new_type:
                    questions.update(question_type=new_type)
                    if version_incremented:
                        messages.success(request, f'Question type changed to {new_type} successfully! Survey version incremented to {survey.version}. Students will need to retake the survey.')
                    else:
                        messages.success(request, f'Question type changed to {new_type} successfully!')
            
            return redirect('edit_survey', survey_id=survey.id)
    else:
        form = QuestionBulkForm()
    
    context = {
        'survey': survey,
        'form': form,
        'profile': profile,
        'active_questions': survey.questions.filter(is_active=True).order_by('order'),
    }
    return render(request, 'myapp/question_bulk_operations.html', context)


@login_required
def survey_settings_management(request, survey_id):
    """Enhanced survey settings management"""
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'teacher':
        messages.error(request, 'Access denied. Teacher access required.')
        return redirect('home')
    
    survey = get_object_or_404(Survey, id=survey_id, created_by=request.user)
    
    if request.method == 'POST':
        form = SurveySettingsForm(request.POST, instance=survey)
        if form.is_valid():
            # Check if content fields changed (not administrative fields)
            content_fields = ['title', 'description']
            content_changed = any(
                form.cleaned_data.get(field) != getattr(survey, field)
                for field in content_fields
            )
            
            # Only increment version if content changed and survey has responses
            if content_changed and survey.responses.exists():
                survey.version += 1
                messages.success(request, f'Survey settings updated successfully! Version incremented to {survey.version}. Students will need to retake the survey.')
            else:
                messages.success(request, 'Survey settings updated successfully!')
            
            form.save()
            return redirect('survey_settings_management', survey_id=survey.id)
    else:
        form = SurveySettingsForm(instance=survey)
    
    # Get survey statistics
    stats = {
        'total_questions': survey.questions.filter(is_active=True).count(),
        'total_responses': survey.responses.count(),
        'assigned_sections': survey.sections.count(),
        'is_open': survey.is_open,
        'days_until_due': None,
    }
    
    if survey.due_date:
        now = timezone.now()
        if survey.due_date > now:
            delta = survey.due_date - now
            stats['days_until_due'] = delta.days
    
    context = {
        'survey': survey,
        'form': form,
        'stats': stats,
        'profile': profile,
    }
    return render(request, 'myapp/survey_settings_management.html', context)


@login_required
def assignment_management(request, survey_id):
    """Manage survey assignments to sections"""
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'teacher':
        messages.error(request, 'Access denied. Teacher access required.')
        return redirect('home')
    
    survey = get_object_or_404(Survey, id=survey_id, created_by=request.user)
    
    if request.method == 'POST':
        form = AssignmentForm(request.POST)
        if form.is_valid():
            sections = form.cleaned_data['sections']
            due_date = form.cleaned_data['due_date']
            is_active = form.cleaned_data['is_active']
            
            # Update survey assignments
            survey.sections.set(sections)
            survey.due_date = due_date
            survey.is_active = is_active
            survey.save()
            
            messages.success(request, 'Survey assignments updated successfully!')
            return redirect('assignment_management', survey_id=survey.id)
    else:
        # Pre-populate form with current survey data
        form = AssignmentForm(initial={
            'sections': survey.sections.all(),
            'due_date': survey.due_date,
            'is_active': survey.is_active,
        })
    
    # Get assignment statistics
    assignment_stats = []
    for section in survey.sections.all():
        students_in_section = UserProfile.objects.filter(section=section, role='student', is_active=True).count()
        responses_from_section = SurveyResponse.objects.filter(
            survey=survey,
            student__userprofile__section=section
        ).count()
        
        assignment_stats.append({
            'section': section,
            'total_students': students_in_section,
            'responses_received': responses_from_section,
            'completion_rate': (responses_from_section / students_in_section * 100) if students_in_section > 0 else 0,
        })
    
    context = {
        'survey': survey,
        'form': form,
        'assignment_stats': assignment_stats,
        'profile': profile,
    }
    return render(request, 'myapp/assignment_management.html', context)


@login_required
def section_bulk_operations(request):
    """Handle bulk operations on sections"""
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'teacher':
        messages.error(request, 'Access denied. Teacher access required.')
        return redirect('home')
    
    if request.method == 'POST':
        form = SectionBulkForm(request.POST)
        if form.is_valid():
            action = form.cleaned_data['action']
            section_ids = form.cleaned_data['section_ids'].split(',')
            sections = Section.objects.filter(id__in=section_ids)
            
            if action == 'delete':
                count = sections.count()
                total_students_deactivated = 0
                
                # Soft delete: mark sections as inactive and deactivate students
                for section in sections:
                    # Get students in this section
                    students_in_section = UserProfile.objects.filter(section=section, role='student', is_active=True)
                    student_count = students_in_section.count()
                    
                    # Note: Students are not deactivated since UserProfile doesn't have is_active field
                    if student_count > 0:
                        total_students_deactivated += student_count
                    
                    # Deactivate section
                    section.is_active = False
                    section.save()
                
                if total_students_deactivated > 0:
                    messages.success(request, f'{count} sections deactivated successfully! {total_students_deactivated} student(s) are in these sections.')
                else:
                    messages.success(request, f'{count} sections deactivated successfully!')
            
            elif action == 'activate':
                # For sections, we might want to activate/deactivate surveys assigned to them
                surveys = Survey.objects.filter(sections__in=sections)
                surveys.update(is_active=True)
                messages.success(request, f'Surveys assigned to selected sections activated!')
            
            elif action == 'deactivate':
                surveys = Survey.objects.filter(sections__in=sections)
                surveys.update(is_active=False)
                messages.success(request, f'Surveys assigned to selected sections deactivated!')
            
            return redirect('manage_sections')
    else:
        form = SectionBulkForm()
    
    sections = Section.objects.filter(is_active=True).annotate(
        student_count=models.Count('userprofile', filter=models.Q(userprofile__role='student'))
    ).order_by('name')
    
    context = {
        'sections': sections,
        'form': form,
        'profile': profile,
    }
    return render(request, 'myapp/section_bulk_operations.html', context)


@login_required
def edit_section(request, section_id):
    """Edit a section"""
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'teacher':
        messages.error(request, 'Access denied. Teacher access required.')
        return redirect('home')
    
    section = get_object_or_404(Section, id=section_id)
    
    if request.method == 'POST':
        form = SectionForm(request.POST, instance=section)
        if form.is_valid():
            form.save()
            messages.success(request, 'Section updated successfully!')
            return redirect('manage_sections')
    else:
        form = SectionForm(instance=section)
    
    # Get section statistics
    stats = {
        'total_students': UserProfile.objects.filter(section=section, role='student', is_active=True).count(),
        'total_surveys': Survey.objects.filter(sections=section).count(),
        'active_surveys': Survey.objects.filter(sections=section, is_active=True).count(),
    }
    
    context = {
        'section': section,
        'form': form,
        'stats': stats,
        'profile': profile,
    }
    return render(request, 'myapp/edit_section.html', context)


@login_required
def delete_section(request, section_id):
    """Soft delete a section (mark as inactive) and deactivate all students in the section"""
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'teacher':
        messages.error(request, 'Access denied. Teacher access required.')
        return redirect('home')
    
    section = get_object_or_404(Section, id=section_id, is_active=True)
    
    # Get all students in this section
    students_in_section = UserProfile.objects.filter(section=section, role='student', is_active=True)
    student_count = students_in_section.count()
    
    # Soft delete: mark section as inactive
    section.is_active = False
    section.save()
    
    # Deactivate all students in this section
    students_in_section.update(is_active=False)
    
    if student_count > 0:
        messages.success(request, f'Section deactivated successfully! {student_count} student(s) have also been deactivated.')
    else:
        messages.success(request, 'Section deactivated successfully!')
    
    return redirect('manage_sections')


@login_required
def restore_section(request, section_id):
    """Restore a soft-deleted section (mark as active) and reactivate all students in the section"""
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'teacher':
        messages.error(request, 'Access denied. Teacher access required.')
        return redirect('home')
    
    section = get_object_or_404(Section, id=section_id, is_active=False)
    
    # Get all students in this section
    students_in_section = UserProfile.objects.filter(section=section, role='student', is_active=True)
    student_count = students_in_section.count()
    
    # Restore: mark section as active
    section.is_active = True
    section.save()
    
    # Reactivate all students in this section
    students_in_section.update(is_active=True)
    
    if student_count > 0:
        messages.success(request, f'Section restored successfully! {student_count} student(s) have also been reactivated.')
    else:
        messages.success(request, 'Section restored successfully!')
    
    return redirect('manage_sections')


@login_required
def manage_students(request):
    """Manage students - view all students with their sections and activity"""
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'teacher':
        messages.error(request, 'Access denied. Teacher access required.')
        return redirect('home')
    
    # Get all students with their profiles (both active and inactive)
    students = UserProfile.objects.filter(role='student').select_related('user', 'section').order_by('section__name', 'user__last_name', 'user__first_name')
    
    # Get section filter options
    sections = Section.objects.filter(is_active=True).order_by('name')
    
    # Apply section filter if provided
    section_filter = request.GET.get('section')
    if section_filter and section_filter != 'all':
        students = students.filter(section_id=section_filter)
    
    # Apply status filter if provided
    status_filter = request.GET.get('status', 'all')
    if status_filter == 'active':
        students = students.filter(is_active=True)
    elif status_filter == 'inactive':
        students = students.filter(is_active=False)
    
    # Get student statistics
    student_stats = []
    for student in students:
        # Count survey responses for this student
        response_count = SurveyResponse.objects.filter(student=student.user).count()
        
        # Get latest response date
        latest_response = SurveyResponse.objects.filter(student=student.user).order_by('-submitted_at').first()
        
        student_stats.append({
            'student': student,
            'response_count': response_count,
            'latest_response': latest_response.submitted_at if latest_response else None,
            'is_active': student.is_active,  # Use UserProfile.is_active instead of user.is_active
        })
    
    # Calculate statistics
    total_students = UserProfile.objects.filter(role='student').count()
    active_students = UserProfile.objects.filter(role='student', is_active=True).count()
    inactive_students = UserProfile.objects.filter(role='student', is_active=False).count()
    
    context = {
        'student_stats': student_stats,
        'sections': sections,
        'current_section_filter': section_filter,
        'current_status_filter': status_filter,
        'profile': profile,
        'total_students': total_students,
        'active_students': active_students,
        'inactive_students': inactive_students,
        'filtered_students': students.count(),
    }
    return render(request, 'myapp/manage_students.html', context)


@login_required
def activate_student(request, student_id):
    """Activate a student (set is_active=True)"""
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'teacher':
        messages.error(request, 'Access denied. Teacher access required.')
        return redirect('home')
    
    student_profile = get_object_or_404(UserProfile, id=student_id, role='student')
    
    if student_profile.is_active:
        messages.warning(request, f'{student_profile.user.get_full_name()} is already active.')
    else:
        student_profile.is_active = True
        student_profile.save()
        messages.success(request, f'{student_profile.user.get_full_name()} has been activated successfully.')
    
    return redirect('manage_students')


@login_required
def deactivate_student(request, student_id):
    """Deactivate a student (set is_active=False)"""
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'teacher':
        messages.error(request, 'Access denied. Teacher access required.')
        return redirect('home')
    
    student_profile = get_object_or_404(UserProfile, id=student_id, role='student')
    
    if not student_profile.is_active:
        messages.warning(request, f'{student_profile.user.get_full_name()} is already inactive.')
    else:
        student_profile.is_active = False
        student_profile.save()
        messages.success(request, f'{student_profile.user.get_full_name()} has been deactivated successfully.')
    
    return redirect('manage_students')


@login_required
def question_reorder(request, survey_id):
    """AJAX endpoint for reordering questions"""
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'teacher':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    survey = get_object_or_404(Survey, id=survey_id, created_by=request.user)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            question_orders = data.get('question_orders', [])
            
            for item in question_orders:
                question_id = item.get('question_id')
                new_order = item.get('order')
                
                question = Question.objects.get(id=question_id, survey=survey)
                question.order = new_order
                question.save()
            
            return JsonResponse({'success': True})
        
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Invalid request method'}, status=405)


@login_required
def analytics_api(request, survey_id):
    """Enhanced AJAX API endpoint for real-time analytics data with filtering"""
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'teacher':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    survey = get_object_or_404(Survey, id=survey_id, created_by=request.user)
    
    if request.method == 'GET':
        try:
            # Get filter parameters
            date_from = request.GET.get('date_from')
            date_to = request.GET.get('date_to')
            section_id = request.GET.get('section_id')
            question_type = request.GET.get('question_type')
            version_filter = request.GET.get('version_filter')
            
            # Apply filters to responses
            responses_query = SurveyResponse.objects.filter(survey=survey)
            
            if date_from:
                from datetime import datetime
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                responses_query = responses_query.filter(submitted_at__date__gte=date_from_obj)
            
            if date_to:
                from datetime import datetime
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                responses_query = responses_query.filter(submitted_at__date__lte=date_to_obj)
            
            if section_id and section_id != 'all':
                responses_query = responses_query.filter(student__userprofile__section_id=section_id)
            
            # Apply version filter
            if version_filter and version_filter != 'all':
                if version_filter == 'current':
                    responses_query = responses_query.filter(survey_version=survey.version)
                elif version_filter == 'outdated':
                    responses_query = responses_query.exclude(survey_version=survey.version)
                elif version_filter == 'latest':
                    # Get only the latest response per student
                    latest_responses = []
                    for student in survey.sections.values_list('userprofile__user', flat=True):
                        latest = responses_query.filter(student_id=student).order_by('-submitted_at').first()
                        if latest:
                            latest_responses.append(latest.id)
                    responses_query = responses_query.filter(id__in=latest_responses)
            
            # Get filtered analytics data
            analytics_data = get_filtered_survey_analytics_data(survey, responses_query, question_type)
            
            # Calculate real-time statistics
            total_responses = responses_query.count()
            total_questions = survey.questions.count()
            
            # Get recent responses (last 24 hours)
            from django.utils import timezone
            from datetime import timedelta
            recent_responses = responses_query.filter(
                submitted_at__gte=timezone.now() - timedelta(hours=24)
            ).count()
            
            # Section completion rates
            section_stats = []
            sections_filter = survey.sections.all()
            if section_id and section_id != 'all':
                sections_filter = sections_filter.filter(id=section_id)
                
            for section in sections_filter:
                students_in_section = UserProfile.objects.filter(section=section, role='student', is_active=True).count()
                responses_from_section = responses_query.filter(
                    student__userprofile__section=section
                ).count()
                
                completion_rate = (responses_from_section / students_in_section * 100) if students_in_section > 0 else 0
                
                section_stats.append({
                    'section_name': section.name,
                    'section_code': section.code,
                    'total_students': students_in_section,
                    'responses_received': responses_from_section,
                    'completion_rate': round(completion_rate, 1),
                })
            
            # Generate timeline data
            timeline_data = generate_response_timeline_data(responses_query)
            
            # Generate insights
            insights = generate_analytics_insights(survey, total_responses, section_stats, analytics_data)
            
            # Calculate version statistics based on filtered responses
            current_version_responses = responses_query.filter(survey_version=survey.version).count()
            outdated_responses = responses_query.exclude(survey_version=survey.version).count()
            total_version_responses = responses_query.count()
            
            version_stats = {
                'current_responses': current_version_responses,
                'outdated_responses': outdated_responses,
                'total_responses': total_version_responses,
                'current_version': survey.version
            }
            
            # Generate version timeline data based on filtered responses
            version_timeline_data = generate_version_timeline_data(survey, responses_query)
            
            return JsonResponse({
                'success': True,
                'data': {
                    'survey_id': survey.id,
                    'survey_title': survey.title,
                    'total_responses': total_responses,
                    'total_questions': total_questions,
                    'recent_responses': recent_responses,
                    'section_stats': section_stats,
                    'analytics_data': analytics_data,
                    'timeline_data': timeline_data,
                    'version_stats': version_stats,
                    'version_data': version_timeline_data,
                    'insights': insights,
                    'filters_applied': {
                        'date_from': date_from,
                        'date_to': date_to,
                        'section_id': section_id,
                        'question_type': question_type,
                        'version_filter': version_filter
                    },
                    'last_updated': timezone.now().isoformat(),
                }
            })
        
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Invalid request method'}, status=405)


def get_dashboard_analytics_data(user):
    """Get comprehensive analytics data for the teacher dashboard"""
    from django.db.models import Count, Q
    from datetime import datetime, timedelta
    
    # Get all surveys by this teacher
    surveys = Survey.objects.filter(created_by=user)
    
    # Pie Chart Data - Response percentages per survey
    pie_chart_data = []
    for survey in surveys:
        total_responses = survey.responses.count()
        total_possible = 0
        
        # Calculate total possible responses (all active students in assigned sections)
        for section in survey.sections.all():
            total_possible += UserProfile.objects.filter(section=section, role='student', is_active=True).count()
        
        response_percentage = (total_responses / total_possible * 100) if total_possible > 0 else 0
        
        pie_chart_data.append({
            'survey_id': survey.id,
            'survey_title': survey.title,
            'responses': total_responses,
            'possible': total_possible,
            'percentage': round(response_percentage, 1)
        })
    
    # Bar Chart Data - Responses per section
    sections = Section.objects.all()
    bar_chart_data = []
    
    for section in sections:
        response_count = SurveyResponse.objects.filter(
            student__userprofile__section=section,
            survey__created_by=user
        ).count()
        
        bar_chart_data.append({
            'section_id': section.id,
            'section_name': section.name,
            'section_code': section.code,
            'response_count': response_count
        })
    
    # Line Chart Data - Daily response trends (last 30 days)
    from django.utils import timezone
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)
    
    line_chart_data = []
    current_date = start_date
    
    while current_date <= end_date:
        daily_responses = SurveyResponse.objects.filter(
            survey__created_by=user,
            submitted_at__date=current_date
        ).count()
        
        line_chart_data.append({
            'date': current_date.strftime('%Y-%m-%d'),
            'date_formatted': current_date.strftime('%m/%d'),
            'response_count': daily_responses
        })
        
        current_date += timedelta(days=1)
    
    # Check if data is empty
    has_data = {
        'surveys': surveys.exists(),
        'responses': SurveyResponse.objects.filter(survey__created_by=user).exists(),
        'sections': sections.exists(),
        'pie_chart': any(item['responses'] > 0 for item in pie_chart_data),
        'bar_chart': any(item['response_count'] > 0 for item in bar_chart_data),
        'line_chart': any(item['response_count'] > 0 for item in line_chart_data)
    }
    
    return {
        'pie_chart_data': pie_chart_data,
        'bar_chart_data': bar_chart_data,
        'line_chart_data': line_chart_data,
        'total_surveys': surveys.count(),
        'total_sections': sections.count(),
        'has_data': has_data
    }


@login_required
def dashboard_analytics_api(request):
    """AJAX API endpoint for dashboard analytics with filtering"""
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'teacher':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    if request.method == 'GET':
        try:
            # Get filter parameters
            survey_id = request.GET.get('survey_id')
            section_id = request.GET.get('section_id')
            date_from = request.GET.get('date_from')
            date_to = request.GET.get('date_to')
            
            # Base queryset for responses
            responses_query = SurveyResponse.objects.filter(survey__created_by=request.user)
            
            # Apply filters
            if survey_id and survey_id != 'all':
                responses_query = responses_query.filter(survey_id=survey_id)
            
            if section_id and section_id != 'all':
                responses_query = responses_query.filter(student__userprofile__section_id=section_id)
            
            if date_from:
                from datetime import datetime
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                responses_query = responses_query.filter(submitted_at__date__gte=date_from_obj)
            
            if date_to:
                from datetime import datetime
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                responses_query = responses_query.filter(submitted_at__date__lte=date_to_obj)
            
            # Get filtered analytics data
            filtered_analytics = get_filtered_dashboard_analytics(request.user, responses_query, survey_id, section_id, date_from, date_to)
            
            return JsonResponse({
                'success': True,
                'data': filtered_analytics
            })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Invalid request method'}, status=405)


def get_filtered_dashboard_analytics(user, responses_query, survey_id=None, section_id=None, date_from=None, date_to=None):
    """Get filtered analytics data based on the provided filters"""
    from django.db.models import Count, Q
    from datetime import datetime, timedelta
    
    # Pie Chart Data - filtered by survey if specified
    surveys = Survey.objects.filter(created_by=user)
    if survey_id and survey_id != 'all':
        surveys = surveys.filter(id=survey_id)
    
    pie_chart_data = []
    for survey in surveys:
        # Filter responses for this survey
        survey_responses = responses_query.filter(survey=survey)
        total_responses = survey_responses.count()
        
        total_possible = 0
        # Calculate total possible responses
        sections_filter = survey.sections.all()
        if section_id and section_id != 'all':
            sections_filter = sections_filter.filter(id=section_id)
            
        for section in sections_filter:
            total_possible += UserProfile.objects.filter(section=section, role='student', is_active=True).count()
        
        response_percentage = (total_responses / total_possible * 100) if total_possible > 0 else 0
        
        pie_chart_data.append({
            'survey_id': survey.id,
            'survey_title': survey.title,
            'responses': total_responses,
            'possible': total_possible,
            'percentage': round(response_percentage, 1)
        })
    
    # Bar Chart Data - responses per section
    sections = Section.objects.all()
    if section_id and section_id != 'all':
        sections = sections.filter(id=section_id)
    
    bar_chart_data = []
    for section in sections:
        section_responses = responses_query.filter(student__userprofile__section=section)
        response_count = section_responses.count()
        
        bar_chart_data.append({
            'section_id': section.id,
            'section_name': section.name,
            'section_code': section.code,
            'response_count': response_count
        })
    
    # Line Chart Data - daily trends within date range
    from django.utils import timezone
    
    if date_from and date_to:
        start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
        end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
    else:
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=30)
    
    line_chart_data = []
    current_date = start_date
    
    while current_date <= end_date:
        daily_responses = responses_query.filter(submitted_at__date=current_date).count()
        
        line_chart_data.append({
            'date': current_date.strftime('%Y-%m-%d'),
            'date_formatted': current_date.strftime('%m/%d'),
            'response_count': daily_responses
        })
        
        current_date += timedelta(days=1)
    
    # Check if filtered data is empty
    has_data = {
        'surveys': surveys.exists(),
        'responses': responses_query.exists(),
        'sections': sections.exists(),
        'pie_chart': any(item['responses'] > 0 for item in pie_chart_data),
        'bar_chart': any(item['response_count'] > 0 for item in bar_chart_data),
        'line_chart': any(item['response_count'] > 0 for item in line_chart_data)
    }
    
    return {
        'pie_chart_data': pie_chart_data,
        'bar_chart_data': bar_chart_data,
        'line_chart_data': line_chart_data,
        'has_data': has_data,
        'filters_applied': {
            'survey_id': survey_id,
            'section_id': section_id,
            'date_from': date_from,
            'date_to': date_to
        }
    }


@login_required
def reorder_questions(request, survey_id):
    """Handle AJAX request to reorder questions"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)
    
    try:
        profile = UserProfile.objects.get(user=request.user)
        if profile.role != 'teacher':
            return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
        
        survey = get_object_or_404(Survey, id=survey_id, created_by=request.user)
        
        # Parse JSON data
        data = json.loads(request.body)
        questions_data = data.get('questions', [])
        
        # Update question orders
        for question_data in questions_data:
            question_id = question_data.get('id')
            new_order = question_data.get('order')
            
            if question_id and new_order:
                question = get_object_or_404(Question, id=question_id, survey=survey)
                question.order = new_order
                question.save()
        
        return JsonResponse({'success': True, 'message': 'Question order updated successfully'})
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# Export functionality for analytics
@login_required
def export_analytics_csv(request, survey_id):
    """Export survey analytics data to CSV"""
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'teacher':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    survey = get_object_or_404(Survey, id=survey_id, created_by=request.user)
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{survey.title}_analytics_{datetime.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    
    # Write header
    writer.writerow(['Survey Analytics Export'])
    writer.writerow(['Survey Title', survey.title])
    writer.writerow(['Export Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    writer.writerow(['Total Responses', survey.responses.count()])
    writer.writerow(['Total Questions', survey.questions.count()])
    writer.writerow([])
    
    # Write section statistics
    writer.writerow(['Section Statistics'])
    writer.writerow(['Section Name', 'Section Code', 'Total Students', 'Responses Received', 'Completion Rate (%)'])
    
    for section in survey.sections.all():
        students_in_section = UserProfile.objects.filter(section=section, role='student', is_active=True).count()
        responses_from_section = SurveyResponse.objects.filter(
            survey=survey,
            student__userprofile__section=section
        ).count()
        completion_rate = (responses_from_section / students_in_section * 100) if students_in_section > 0 else 0
        
        writer.writerow([
            section.name,
            section.code,
            students_in_section,
            responses_from_section,
            round(completion_rate, 1)
        ])
    
    writer.writerow([])
    
    # Write question analytics
    writer.writerow(['Question Analytics'])
    writer.writerow(['Question Number', 'Question Text', 'Question Type', 'Is Required', 'Response Count', 'Analytics Data'])
    
    analytics_data = get_survey_analytics_data(survey)
    for i, data in enumerate(analytics_data, 1):
        question = data['question']
        response_count = Answer.objects.filter(response__survey=survey, question=question).count()
        
        # Format analytics data
        analytics_text = ""
        if data['stats']:
            analytics_text = "; ".join([f"{choice}: {stats['count']} ({stats['percentage']}%)" 
                                       for choice, stats in data['stats'].items()])
        
        writer.writerow([
            i,
            question.question_text,
            question.get_question_type_display(),
            'Yes' if question.is_required else 'No',
            response_count,
            analytics_text
        ])
    
    return response


@login_required
def export_analytics_excel(request, survey_id):
    """Export survey analytics data to Excel format"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return JsonResponse({'error': 'Excel export requires openpyxl package'}, status=500)
    
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'teacher':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    survey = get_object_or_404(Survey, id=survey_id, created_by=request.user)
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Survey Analytics"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write header information
    ws['A1'] = f"Analytics Report: {survey.title}"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A3'] = f"Total Responses: {survey.responses.count()}"
    ws['A4'] = f"Total Questions: {survey.questions.count()}"
    
    # Write section statistics
    row = 6
    ws[f'A{row}'] = "Section Statistics"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    
    # Section headers
    headers = ['Section Name', 'Section Code', 'Total Students', 'Responses Received', 'Completion Rate (%)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    row += 1
    
    # Section data
    for section in survey.sections.all():
        students_in_section = UserProfile.objects.filter(section=section, role='student', is_active=True).count()
        responses_from_section = SurveyResponse.objects.filter(
            survey=survey,
            student__userprofile__section=section
        ).count()
        completion_rate = (responses_from_section / students_in_section * 100) if students_in_section > 0 else 0
        
        ws.cell(row=row, column=1, value=section.name)
        ws.cell(row=row, column=2, value=section.code)
        ws.cell(row=row, column=3, value=students_in_section)
        ws.cell(row=row, column=4, value=responses_from_section)
        ws.cell(row=row, column=5, value=round(completion_rate, 1))
        row += 1
    
    # Write question analytics
    row += 2
    ws[f'A{row}'] = "Question Analytics"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    
    # Question headers
    q_headers = ['Question #', 'Question Text', 'Type', 'Required', 'Response Count', 'Analytics']
    for col, header in enumerate(q_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    row += 1
    
    # Question data
    analytics_data = get_survey_analytics_data(survey)
    for i, data in enumerate(analytics_data, 1):
        question = data['question']
        response_count = Answer.objects.filter(response__survey=survey, question=question).count()
        
        # Format analytics data
        analytics_text = ""
        if data['stats']:
            analytics_text = "; ".join([f"{choice}: {stats['count']} ({stats['percentage']}%)" 
                                       for choice, stats in data['stats'].items()])
        
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=question.question_text)
        ws.cell(row=row, column=3, value=question.get_question_type_display())
        ws.cell(row=row, column=4, value='Yes' if question.is_required else 'No')
        ws.cell(row=row, column=5, value=response_count)
        ws.cell(row=row, column=6, value=analytics_text)
        row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{survey.title}_analytics_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response


@login_required
def export_responses_csv(request, survey_id):
    """Export individual survey responses to CSV"""
    profile = UserProfile.objects.get(user=request.user)
    if profile.role != 'teacher':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    survey = get_object_or_404(Survey, id=survey_id, created_by=request.user)
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{survey.title}_responses_{datetime.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    
    # Get all responses
    responses = SurveyResponse.objects.filter(survey=survey).order_by('-submitted_at')
    
    # Create header row
    header = ['Student Name', 'Student Username', 'Section', 'Submitted At', 'Survey Version']
    
    # Add question columns
    questions = survey.questions.all().order_by('order')
    for question in questions:
        header.append(f"Q{question.order}: {question.question_text[:50]}...")
    
    writer.writerow(header)
    
    # Write response data
    for response_obj in responses:
        try:
            student_profile = UserProfile.objects.get(user=response_obj.student)
            section_name = student_profile.section.name if student_profile.section else 'No Section'
        except UserProfile.DoesNotExist:
            section_name = 'No Profile'
        
        row = [
            response_obj.student.get_full_name() or response_obj.student.username,
            response_obj.student.username,
            section_name,
            response_obj.submitted_at.strftime('%Y-%m-%d %H:%M:%S'),
            response_obj.survey_version
        ]
        
        # Add answers for each question
        for question in questions:
            try:
                answer = Answer.objects.get(response=response_obj, question=question)
                if question.question_type in ['multiple_choice', 'likert_scale']:
                    row.append(answer.answer_choice)
                else:
                    row.append(answer.answer_text)
            except Answer.DoesNotExist:
                row.append('No Answer')
        
        writer.writerow(row)
    
    return response

